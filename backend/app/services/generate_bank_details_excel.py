"""Generate an Excel file with student banking details for employment documents."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.styles.borders import Border
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.bank_details import BankDetails
from app.models.student import Student
from app.services.fio_inflection import fio_to_dative


HEADERS = [
    "№ договора",
    "Дата",
    "Срок договора",
    "Срок оказ. Услуг",
    "ФИО",
    "Фамилия",
    "Инициалы",
    "ФИО в Д.П.",
    "Фамилия в Д.Ф.",
    "Сумма",
    "Сумм. словами Р.П,",
    "Дата рождения",
    "СНИЛС",
    "ИНН",
    "Контактный телефон",
    "Адрес электронной корпоративной почты",
    "Серия паспорта",
    "Номер паспорта",
    "Кем выдан паспорт",
    "Дата выдачи паспорта",
    'Место регистрации по паспорту ("Прописка")',
    "Место фактического проживания",
    "Номер счета",
    "Банк получателя: ",
    "БИК",
    "Корр. счет",
    "Заказчик",
    "Его статус ",
    "Тел.",
    "Почта",
]

TEXT_COLUMNS = {1, 2, 3, 4, 12, 13, 14, 15, 17, 18, 20, 23, 25, 26, 29}
DEFAULT_COLUMN_WIDTHS = {
    1: 12,
    2: 18,
    3: 24,
    4: 24,
    5: 32,
    6: 18,
    7: 12,
    8: 34,
    9: 20,
    10: 12,
    11: 24,
    12: 14,
    13: 18,
    14: 18,
    15: 18,
    16: 34,
    17: 14,
    18: 14,
    19: 36,
    20: 16,
    21: 44,
    22: 44,
    23: 24,
    24: 24,
    25: 14,
    26: 24,
    27: 28,
    28: 36,
    29: 18,
    30: 28,
}


def _student_full_name(student: Student) -> str:
    return " ".join(part for part in (student.last_name, student.first_name, student.middle_name) if part)


def _digits_only(value: str | None, *, max_len: int | None = None) -> str:
    digits = re.sub(r"\D", "", value or "")
    if max_len is not None:
        return digits[:max_len]
    return digits


def _format_date(value: date | datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y")


def _format_snils(value: str | None) -> str:
    digits = _digits_only(value, max_len=11)
    if len(digits) != 11:
        return value or ""
    return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]} {digits[9:]}"


def _format_inn(value: str | None) -> str:
    digits = _digits_only(value, max_len=12)
    if len(digits) != 12:
        return value or ""
    return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}"


def _format_phone(value: str | None) -> str:
    digits = _digits_only(value, max_len=11)
    if len(digits) == 11 and digits.startswith(("7", "8")):
        return f"8({digits[1:4]}){digits[4:7]}-{digits[7:9]}-{digits[9:]}"
    if len(digits) == 10:
        return f"8({digits[:3]}){digits[3:6]}-{digits[6:8]}-{digits[8:]}"
    return value or ""


def _initials(student: Student) -> str:
    parts = [student.first_name, student.middle_name]
    return " ".join(f"{part.strip()[0]}." for part in parts if part and part.strip())


def _pick_active_bank_details(rows: list[BankDetails]) -> BankDetails | None:
    active_rows = [row for row in rows if row.is_active]
    candidates = active_rows or rows
    if not candidates:
        return None
    return max(candidates, key=lambda row: row.bank_details_id)


def _missing_fields(student: Student, bank_details: BankDetails | None) -> list[str]:
    missing: list[str] = []
    if not _student_full_name(student):
        missing.append("ФИО")
    if student.birth_date is None:
        missing.append("Дата рождения")
    if not _digits_only(student.snils, max_len=11):
        missing.append("СНИЛС")
    if not _digits_only(student.inn, max_len=12):
        missing.append("ИНН")
    if not (student.phone or "").strip():
        missing.append("Телефон")
    if not (student.corporate_email or student.email or "").strip():
        missing.append("Email")
    if not _digits_only(student.passport_series, max_len=4):
        missing.append("Серия паспорта")
    if not _digits_only(student.passport_number, max_len=6):
        missing.append("Номер паспорта")
    if not (student.passport_issued_by or "").strip():
        missing.append("Кем выдан паспорт")
    if student.passport_issue_date is None:
        missing.append("Дата выдачи паспорта")
    if not (student.registration_address or "").strip():
        missing.append("Место регистрации")
    if bank_details is None:
        missing.append("Банковские реквизиты")
    else:
        if not (bank_details.bank_name or "").strip():
            missing.append("Банк получателя")
        if not _digits_only(bank_details.bik, max_len=9):
            missing.append("БИК")
        if not _digits_only(bank_details.account_number, max_len=20):
            missing.append("Номер счета")
    return missing


def _student_row(student: Student, bank_details: BankDetails) -> list[str | int]:
    registration_address = (student.registration_address or "").strip()
    residential_address = (student.residential_address or "").strip() or registration_address

    return [
        "",
        "",
        "",
        "",
        _student_full_name(student),
        student.last_name,
        _initials(student),
        fio_to_dative(student.last_name, student.first_name, student.middle_name),
        fio_to_dative(student.last_name, "", "").strip() or student.last_name,
        "",
        "",
        _format_date(student.birth_date),
        _format_snils(student.snils),
        _format_inn(student.inn),
        _format_phone(student.phone),
        (student.corporate_email or student.email or "").strip(),
        _digits_only(student.passport_series, max_len=4),
        _digits_only(student.passport_number, max_len=6),
        (student.passport_issued_by or "").strip(),
        _format_date(student.passport_issue_date),
        registration_address,
        residential_address,
        _digits_only(bank_details.account_number, max_len=20),
        (bank_details.bank_name or "").strip(),
        _digits_only(bank_details.bik, max_len=9),
        _digits_only(bank_details.correspondent_account, max_len=23),
        "",
        "",
        "",
        "",
    ]


async def build_bank_details_rows(session: AsyncSession, student_ids: list[int]) -> list[list[str | int]]:
    unique_ids = list(dict.fromkeys(student_ids))
    if not unique_ids:
        raise HTTPException(status_code=422, detail="Выберите участников для формирования Excel файла.")

    stmt = (
        select(Student)
        .options(selectinload(Student.bank_details_rows))
        .where(Student.student_id.in_(unique_ids))
    )
    result = await session.execute(stmt)
    students_by_id = {student.student_id: student for student in result.scalars().all()}

    missing_students = [student_id for student_id in unique_ids if student_id not in students_by_id]
    if missing_students:
        raise HTTPException(status_code=404, detail=f"Студенты не найдены: {', '.join(map(str, missing_students))}.")

    rows: list[list[str | int]] = []
    for student_id in unique_ids:
        student = students_by_id[student_id]
        bank_details = _pick_active_bank_details(list(student.bank_details_rows))
        missing = _missing_fields(student, bank_details)
        if missing:
            raise HTTPException(
                status_code=422,
                detail=f"Для студента {_student_full_name(student)} не заполнены поля: {', '.join(missing)}.",
            )
        assert bank_details is not None
        rows.append(_student_row(student, bank_details))

    return rows


def generate_bank_details_excel(rows: list[list[str | int]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ответы на форму (1)"

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.column in TEXT_COLUMNS:
                cell.number_format = "@"

    for column_index, width in DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 36

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def make_bank_details_excel_filename() -> str:
    return "Заполнение_данных_банк_реквизиты.xlsx"
